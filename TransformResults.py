from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
import os
import numpy as np
import xlsxwriter

Path="G:\\MASTER_THESIS\\DanielKOD\\CompletCode\\Results\\"
PFileRF=Path+"RF3Diff"
PFileSEA=Path+"SEA"

workbook = xlsxwriter.Workbook(Path+'Results.xlsx')
worksheet = workbook.add_worksheet()

for i in range(1,12):
    ii=str(i)
    
    read_file = pd.read_csv (r''+Path+'RF3Diff'+ii+'.csv')
    read_file.to_excel (r''+Path+'RF3Diff'+ii+'.xlsx', index = None, header=True)

    read_file = pd.read_csv (r''+Path+'SEA'+ii+'.csv')
    read_file.to_excel (r''+Path+'SEA'+ii+'.xlsx', index = None, header=True)
    
   
    wb= load_workbook(PFileRF+ii+'.xlsx')
    ws= wb.active

    row_count = ws.max_row
    row_count=str(row_count)
    

    df = pd.read_excel(PFileRF+ii+'.xlsx')
    MaxValueRF=df['0'].max()
    
    df = pd.read_excel(PFileSEA+ii+'.xlsx')
    MaxValueSEA=df['0.000000'].max()
    
    row=1+i
    row=str(row)
    worksheet.write('A'+row, 'sim'+ii)
    worksheet.write('B1', 'SEA')
    worksheet.write('C1', 'Diff3Max')
    worksheet.write('B'+row, MaxValueSEA)
    worksheet.write('C'+row, MaxValueRF)

    os.remove(''+Path+'RF3Diff'+ii+'.csv')
    os.remove(''+Path+'SEA'+ii+'.csv')
    os.remove(''+Path+'RF3Diff'+ii+'.xlsx')
    os.remove(''+Path+'SEA'+ii+'.xlsx')
    
workbook.close()

